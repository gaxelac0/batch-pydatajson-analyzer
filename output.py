#%matplotlib inline
from asyncio import constants
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from datetime import date
import pandas as pd
import constants
from datetime import datetime
from typing import Counter

from pydatajson import helpers, readers

dfc = 'distribuciones_formatos_cant'


def generate_dataset_error_qty_pie_chart(stats_counter, workbook_filename):

    result_counter = stats_counter
    result_counter.update({"dataset_ok": result_counter['datasets_cant']-result_counter['datasets_meta_error_cant']})

    labels = ['dataset_ok', 'datasets_meta_error_cant']

    values = []
    for label in labels: 
        values.append(result_counter[label])

    fig = plt.figure(figsize=(10,7))
    plt.pie(values, labels = labels, autopct='%1.1f%%', pctdistance=1.1, labeldistance=1.2)

    name_image = './/test//result//charts//dataset_error_qty_pie_chart.png'

    plt.savefig(name_image)

    # se inserta el grafico creado en la worksheet indicada en CHART_SHEET_NAME
    import openpyxl
    wb = openpyxl.load_workbook(filename = workbook_filename)
    ws = wb[constants.CHART_SHEET_NAME]
    img = openpyxl.drawing.image.Image(name_image)
    ws.cell(row=1, column=13, value="Chartpie de % de dataset erroneos")
    img.anchor = 'M2'
    ws.add_image(img)
    wb.save(workbook_filename)


def generate_distribution_types_pie_chart(stats_counter, qty, workbook_filename):

    all = stats_counter.most_common(None)

    qty = len(all) if qty > len(all) else qty

    most_common = all[:qty]
    others = all[qty:]

    others_val = 0
    others_legend = []
    for tuple in others:
        others_legend.append(tuple[0])
        others_val += tuple[1]

    #others_tuple = (', '.join(others_legend), others_val)
    others_tuple = ('OTHERS', others_val)

    result_counter = most_common
    result_counter.append(others_tuple)

    labels = []
    for tuple in result_counter: 
        labels.append(tuple[0])

    values = []
    for tuple in result_counter: 
        values.append(tuple[1])

    explode = []
    for i in range(qty):
        explode.append(0)
    explode.append(0.3)

    fig = plt.figure(figsize=(10,7))
    plt.pie(values, labels = labels, autopct='%1.1f%%', pctdistance=1.1, labeldistance=1.2, explode=explode)

    others_patch = mpatches.Patch(color='gray', label="OTHERS: "+', '.join(others_legend))

    plt.legend(handles=[others_patch])

    name_image = './/test//result//charts//distribution-types-pie-chart-'+str(qty)+'.png'

    plt.savefig(name_image)

    # se inserta el grafico creado en la worksheet indicada en CHART_SHEET_NAME
    import openpyxl
    wb = openpyxl.load_workbook(filename = workbook_filename)
    ws = wb.create_sheet(constants.CHART_SHEET_NAME)
    ws.cell(row=1, column=1, value="Chartpie de % de tipos de distribuciones")
    img = openpyxl.drawing.image.Image(name_image)
    img.anchor = 'A2'
    ws.add_image(img)
    wb.save(workbook_filename)



def generate_dataset_accrual_periodicity_chartpie(accrual_periodicity_counter, workbook_filename):

    result_counter = accrual_periodicity_counter.most_common(None)

    values = []
    labels = []

    for elem in result_counter: 
        labels.append(elem[0])
        values.append(elem[1])

    fig = plt.figure(figsize=(10,7))
    plt.pie(values, labels = labels, autopct='%1.1f%%', pctdistance=1.1, labeldistance=1.2)

    name_image = './/test//result//charts//dataset_accrual_periodicity_chartpie.png'

    plt.savefig(name_image)

    # se inserta el grafico creado en la worksheet indicada en CHART_SHEET_NAME
    import openpyxl
    wb = openpyxl.load_workbook(filename = workbook_filename)
    ws = wb[constants.CHART_SHEET_NAME]
    img = openpyxl.drawing.image.Image(name_image)
    ws.cell(row=1, column=23, value="Chartpie de % de accrualPeriodicity")
    img.anchor = 'Z2'
    ws.add_image(img)
    wb.save(workbook_filename)

import matplotlib.pyplot as plt
import numpy as np
def generate_bar_graph_outdated_datasets(accrual_periodicity_counter, not_updated_dataset_counter, workbook_filename):
  
    # ordenamos los counters, eliminamos nulos, 
    accrual_periodicity_counter = accrual_periodicity_counter.most_common(None)
    not_updated_dataset_counter = not_updated_dataset_counter.most_common(None)

    # generamos arrays utiles para calculo
    totals_label = []
    totals_value = []
    not_updated_label = []
    not_updated_value = []
    for elem in accrual_periodicity_counter: 
        totals_label.append(elem[0])
        totals_value.append(elem[1])

    for elem in not_updated_dataset_counter: 
        not_updated_label.append(elem[0])
        not_updated_value.append(elem[1])


    # numpy arrays calculation
    updated_np = []
    outdated_np = []

    labels = []

    for idx, label in enumerate(totals_label):

        labels.append(label)
        if label in not_updated_label:
            updated_np.append(totals_value[idx]-not_updated_value[not_updated_label.index(label)])
            outdated_np.append(not_updated_value[not_updated_label.index(label)])
        else: 
            updated_np.append(totals_value[idx])
            outdated_np.append(0)


    status_counts = {
        'Updated': np.array(updated_np),
        'Outdated': np.array(outdated_np),
    }
    width = 0.6  # the width of the bars: can also be len(x) sequence


    fig, ax = plt.subplots()
    bottom = np.zeros(len(totals_label))

    for status, status_count in status_counts.items():
        p = ax.bar(labels, status_count, width, label=status, bottom=bottom)
        bottom += status_count

        ax.bar_label(p, label_type='center')

    ax.set_title('Number of outdated dataset per accrual periodicity')
    ax.legend()

    #plt.show()

    name_image = './/test//result//charts//generate_bar_graph_outdated_datasets.png'

    plt.savefig(name_image)

        # se inserta el grafico creado en la worksheet indicada en CHART_SHEET_NAME
    import openpyxl
    wb = openpyxl.load_workbook(filename = workbook_filename)
    ws = wb[constants.CHART_SHEET_NAME]
    img = openpyxl.drawing.image.Image(name_image)
    ws.cell(row=1, column=23, value="Bar Label de Actualizacion de Accrual Periodicy")
    img.anchor = 'AL2'
    ws.add_image(img)
    wb.save(workbook_filename)


def generate_output_catalog_indicator(curr_dataframe, source, accesible, indicators=None, dist_type_counter=None):

    if not isinstance(curr_dataframe, pd.DataFrame) and curr_dataframe == None:
        curr_dataframe = write_catalog_indicator_header()
   
    if indicators != None:
        indicators = indicators[0][0]

        # actualizamos el contador de estadisticas de tipos de distribucion 
        dist_type_counter.update(indicators[dfc])

        append_dataframe = pd.DataFrame({
            'Origen/URL': [source],
            'Accesible ('+ date.today().strftime("%d/%m/%Y") +')': ['Si' if accesible else 'No'],
            'Titulo': [indicators['title']],
            'Actualizado hace (dias)': [indicators['catalogo_ultima_actualizacion_dias']],
            'Cantidad Dataset': [indicators['datasets_cant']],
            'Cantidad Dataset Erroneos': [indicators['datasets_meta_error_cant']],
            'JSON': [indicators[dfc]['JSON'] if 'JSON' in indicators[dfc] else 0],
            'PDF': [indicators[dfc]['PDF'] if 'PDF' in indicators[dfc] else 0],
            'XLS': [indicators[dfc]['XLS'] if 'XLS' in indicators[dfc] else 0],
            'CSV': [indicators[dfc]['CSV'] if 'CSV' in indicators[dfc] else 0],
            'JPG': [indicators[dfc]['JPG'] if 'JPG' in indicators[dfc] else 0],
            'XML': [indicators[dfc]['XML'] if 'XML' in indicators[dfc] else 0],
            'DOC': [indicators[dfc]['DOC'] if 'DOC' in indicators[dfc] else 0],
            'PPT': [indicators[dfc]['PPT'] if 'PPT' in indicators[dfc] else 0],
        })
    return pd.concat([curr_dataframe, append_dataframe], axis=0)


def generate_output_per_dataset(curr_dataframe, source, result, validation_report, catalog_dict, accrual_periodicity_counter, not_updated_dataset_counter):

    if not isinstance(curr_dataframe, pd.DataFrame) and curr_dataframe == None:
        curr_dataframe = write_per_dataset_header()

    categories = validation_report['error']
    if result == True:
        for cat in categories: 
            if cat == 'dataset' and categories[cat] != None:
                for dataset in categories[cat]:

                    append_dataframe = pd.DataFrame({ 
                        'URL': [source],
                        'Catalog Errors': ["No"],
                        'Catalog Error Desc': ['None'],
                        'Dataset title': dataset['title'],
                        'Dataset Errors': ["No"],
                        'Dataset Error Desc': ['None'],
                    })

                    update_accrual_periodicity_counters(catalog_dict, dataset['title'], accrual_periodicity_counter, not_updated_dataset_counter)

                    curr_dataframe = pd.concat([curr_dataframe, append_dataframe], axis=0)
                    #util.write_to_file(file_writer, dataframe=dataframe, sheet_name=constants.DATASET_REPORT_SHEET_NAME)
    else:
        catalog_error = False
        catalog_error_desc = ''
        for cat in categories: 
            if cat == 'catalog' and categories[cat] != None:
                if categories[cat]['status'] != 'OK':
                    catalog_error = True
                    for cat_err in categories[cat]['errors']: 
                        catalog_error_desc = catalog_error_desc + cat_err['message'] +' && '
            elif cat == 'dataset' and categories[cat] != None:
                for dataset in categories[cat]:
                    if dataset['status'] == 'ERROR':
                        final_error_conc = ''
                        for error in dataset['errors']:
                            final_error_conc  = final_error_conc + error['message'] + ' && '
                       
                        append_dataframe = pd.DataFrame({ 
                            'URL': [source],
                            'Catalog Errors': [ 'Si' if catalog_error == True else 'No'],
                            'Catalog Error Desc': [catalog_error_desc],
                            'Dataset title': dataset['title'],
                            'Dataset Errors': ["Si"],
                            'Dataset Error Desc': [final_error_conc],
                        })

                        update_accrual_periodicity_counters(catalog_dict, dataset['title'], accrual_periodicity_counter, not_updated_dataset_counter)

                        curr_dataframe = pd.concat([curr_dataframe, append_dataframe], axis=0)
                       #util.write_to_file(file_writer, dataframe, constants.DATASET_REPORT_SHEET_NAME)
    return curr_dataframe

def write_per_dataset_header():

    # Header Dataframe
    header_df = pd.DataFrame({
         'URL': [],
         'Catalog Errors': [],
         'Catalog Error Desc': [],
         'Dataset title': [],
         'Dataset Errors': [],
         'Dataset Error Desc': [],
    })
    return header_df

def write_catalog_indicator_header():

    header_df = pd.DataFrame({ 
        'Origen/URL': [],
        'Accesible ('+ date.today().strftime("%d/%m/%Y") +')': [],
        'Titulo': [],
        'Actualizado hace (dias)': [],
        'Cantidad Dataset': [],
        'Cantidad Dataset Erroneos': [],
        'JSON': [],
        'PDF': [],
        'XLS': [],
        'CSV': [],
        'JPG': [],
        'XML': [],
        'DOC': [],
        'PPT': [],
    })
    return header_df

# based on DataJson.dataset_is_updated
def update_accrual_periodicity_counters(catalog, dataset, accrual_periodicity_counter, not_updated_dataset_counter):

    catalog = readers.read_catalog(catalog)

    for catalog_dataset in catalog.get('dataset', []):
        if catalog_dataset.get('title') == dataset:
            periodicity = catalog_dataset.get('accrualPeriodicity')
            if not periodicity:
                continue

            if periodicity.casefold() == 'eventual'.casefold():
                continue

            if "modified" not in catalog_dataset:
                accrual_periodicity_counter.update({periodicity: 1})
                not_updated_dataset_counter.update({periodicity: 1})
                continue

            # counter of periodicity
            date = helpers.parse_date_string(catalog_dataset['modified'])
            days_diff = float((datetime.now() - date).days)
            interval = helpers.parse_repeating_time_interval(periodicity)

            
            if days_diff < interval:
                accrual_periodicity_counter.update({periodicity: 1})
                continue

            # if we reach here it means its outdated
            accrual_periodicity_counter.update({periodicity: 1})
            not_updated_dataset_counter.update({periodicity: 1})
